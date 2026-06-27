"""
Agent智能风险监控管理系统 - 数据服务
处理数据导入导出（Excel/DOCX/TXT）
"""

import os
import json
import logging
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


class DataService:
    """数据导入导出服务"""

    BASE_DIR = os.path.dirname(os.path.dirname(__file__))

    @staticmethod
    def import_from_excel(filepath: str) -> List[Dict[str, Any]]:
        """从Excel导入风险数据"""
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl未安装")
            return []

        risks = []
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 假设第一行是表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳过空行
                continue

            risk = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    field_map = {
                        '风险编号': 'risk_code',
                        '风险名称': 'name',
                        '类别': 'category',
                        '子类别': 'subcategory',
                        '风险描述': 'description',
                        '可能后果': 'possible_consequences',
                        '信息来源': 'information_source',
                        '可能性(P)': 'probability',
                        '影响(I)': 'impact',
                        '当前应对': 'current_response',
                        '改进建议': 'suggested_improvement',
                    }
                    key = field_map.get(str(header), str(header).lower().replace(' ', '_'))
                    risk[key] = row[i]

            if risk.get('name'):
                risks.append(risk)

        logger.info(f"从Excel导入 {len(risks)} 个风险")
        return risks

    @staticmethod
    def import_from_docx(filepath: str) -> str:
        """从Word文档提取文本"""
        try:
            from docx import Document
        except ImportError:
            logger.error("python-docx未安装")
            return ""

        doc = Document(filepath)
        paragraphs = []
        for p in doc.paragraphs:
            if p.text.strip():
                paragraphs.append(p.text.strip())

        text = '\n'.join(paragraphs)
        logger.info(f"从Word导入文本: {len(text)} 字符")
        return text

    @staticmethod
    def import_from_txt(filepath: str) -> str:
        """从TXT文件读取文本"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                text = f.read()
            logger.info(f"从TXT导入文本: {len(text)} 字符")
            return text
        except UnicodeDecodeError:
            with open(filepath, 'r', encoding='gbk') as f:
                text = f.read()
            logger.info(f"从TXT导入文本(GBK): {len(text)} 字符")
            return text

    @staticmethod
    def export_to_json(data: Any, filepath: str) -> str:
        """导出为JSON文件"""
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"导出JSON: {filepath}")
        return filepath

    @staticmethod
    def import_from_json(filepath: str) -> Any:
        """从JSON文件导入"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"从JSON导入数据")
        return data

    @staticmethod
    def extract_risk_sections(text: str) -> Dict[str, str]:
        """
        从文档文本中提取风险相关章节
        用于处理年报等长文档
        """
        sections = {
            "risk_factors": "",
            "risk_management": "",
            "internal_control": "",
            "other": ""
        }

        lines = text.split('\n')
        current_section = "other"
        risk_keywords = ['风险', 'risk', '不确定']
        control_keywords = ['内控', '控制', 'control', '治理', '合规']
        management_keywords = ['管理', 'management', '应对', '监督', '审计']

        for line in lines:
            if any(kw in line.lower() for kw in risk_keywords):
                current_section = "risk_factors"
            elif any(kw in line.lower() for kw in management_keywords):
                current_section = "risk_management"
            elif any(kw in line.lower() for kw in control_keywords):
                current_section = "internal_control"

            sections[current_section] += line + '\n'

        return sections

    @staticmethod
    def load_sample_data(enterprise: str = "tencent") -> Optional[Dict[str, Any]]:
        """加载预置的样本数据"""
        data_dir = os.path.join(DataService.BASE_DIR, 'data')

        if enterprise.lower() == "tencent":
            filepath = os.path.join(data_dir, 'tencent_data.json')
            if os.path.exists(filepath):
                return DataService.import_from_json(filepath)

        # 返回通用数据
        from agents.knowledge_base import get_knowledge_base
        kb = get_knowledge_base()

        return {
            "enterprise_name": "腾讯控股",
            "risks": kb.get("tencent_risks", []),
            "internal_controls": kb.get("internal_controls", []),
            "risk_categories": kb.get("risk_categories", {}),
            "response_4t": kb.get("response_4t", {}),
        }
